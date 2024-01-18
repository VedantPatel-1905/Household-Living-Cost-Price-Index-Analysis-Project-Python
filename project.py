import openpyxl,datetime
from openpyxl.styles import Font
import pyinputplus as pyip

# using openpyxl to open excel data sheet and exceptional handling is used to handel any errors
try:
    excelPath = 'household-living-costs-price-indexes-september-2023-quarter-time-series-indexes.xlsx'
    wb = openpyxl.load_workbook(excelPath)
    sheet = wb.active
except FileNotFoundError:
    print("File you are trying to open does not exist")

# This function is use to create new excel file with name based on report type
def createFile(start,end,householdType,group,subgroup):
    if(start != end):
        if(householdType == False):
            path = f'{group}-{subgroup}-{start}-{end}.xlsx'
        elif(group == False):
            path = f'{householdType}-{start}-{end}.xlsx'
        else:
            path = f'{householdType}-{group}-{subgroup}-{start}-{end}.xlsx'
    else:
        if(householdType == False):
            path = f'{group}-{subgroup}-{start}.xlsx'
        elif(group == False):
            path = f'{householdType}-{start}.xlsx'
        else:
            path = f'{householdType}-{group}-{subgroup}-{start}.xlsx'

    openpyxl.Workbook().save(path)
    return path

# This Function is use to enter common column headings to the reports
def enterReportHeadings(newSheet):
    # This loop enter the headings of the data with bold letters into new excel file
    for k in range(1,12):
        c1 = sheet.cell(row=1, column=k)
        data = c1.value
        newSheet.cell(row=11, column=k).value = data
        newSheet.cell(row=11,column=k).font = Font(bold= True)

    # using datetime we get current time and date information to enter it into data
    currentDate = datetime.datetime.now()
    dateTimeValue = f'Report Generation Date and Time\nYear: {currentDate.year}\nMonth: {currentDate.strftime("%B")} \nDate: {currentDate.day}\nTime: {currentDate.strftime("%H:%M:%S")} '
    newSheet.cell(row=3, column=1).value = dateTimeValue

    # This functions are use to style and merge the cells
    newSheet.cell(row=1,column=1).font = Font(size = 26,bold= True)
    newSheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=13 )

    for i in range(3,10):
        newSheet.merge_cells(start_row=i, start_column=1, end_row=i, end_column=13 )
    

# This Function take user input to make report base on a household types
def selectIncomeType():
    # Taking user input to select house hold types
    options = ['All households','Beneficiary','Expenditure quintile 1 (low)','Expenditure quintile 2','Expenditure quintile 3','Expenditure quintile 4','Expenditure quintile 5 (high)','Income quintile 1 (low)','Income quintile 2','Income quintile 3','Income quintile 4','Income quintile 5 (high)','Maori','Super annuitant']
    householdType = pyip.inputMenu(options,prompt="Select the Household type:\n", numbered = True)
    
    return householdType

# This function is used to set the data and generate report in new excel file based on year and household type
def setIncomeTypeReport():  
    # calling function selectYear to get year range
    start,end = selectYear()
    householdType = selectIncomeType()

    # Creating excel report sheet based on year and house hold type
    path = createFile(start,end,householdType,False,False)

    check = sheet['A']
    newWorkBook = openpyxl.load_workbook(path)
    newSheet = newWorkBook.active
    rowNewFile = 11
    avg = change = 0
    avgCount= changeCount = 0
    indexChangeList = []
    costChangeList = []

    # This sets the heading of report base on the categories of report user wants
    if start == end:
        newSheet.cell(row=1,column=1).value = f"Report on Expenditure by {householdType} of year {end}"
    else:
        newSheet.cell(row=1,column=1).value = f"Report on Expenditure by {householdType} from year {start} to {end}"

    # calling this function to add report headings
    enterReportHeadings(newSheet)

    # This loop enters data into new excel file based on year and household type
    j = 1
    for i in check:
        # this is to skip the first iteration which adds column name
        if j != 1:
            if(sheet.cell(row = j,column = 3).value == None):
                break
            else:
                # this handles error if string can't be converted to integer
                try:
                    year = int(sheet.cell(row = j,column = 3).value[0:4])
                except ValueError:
                    print("Cannot convert the value to integer")


            if i.value == householdType and  year >= start and year <= end:
                # this handles error if string can't be converted to integer
                try:
                    # this is to get average money spend by household
                    avg += int(sheet.cell(row = j, column = 9).value)
                except ValueError:
                    print("Cannot convert the value to integer")
                avgCount+=1
                rowNewFile+=1
                # This adds purchase cost into list
                costChange = sheet.cell(row= j, column=9).value
                costChangeList.append(int(costChange))
                # this adds the filtered values to the new excel sheet

                for k in range(1,12):
                    c1 = sheet.cell(row=j, column=k)
                    newSheet.cell(row=rowNewFile, column=k).value = c1.value

                temp = sheet.cell(row = j, column = 11).value
                if(temp != None and temp != "NA"):
                    # This adds purchase change index into list
                    indexChangeList.append(int(temp))
                    change += int(temp)
                    changeCount += 1
        j+=1

    # this is to prevent zero division error
    if avgCount == 0 : avg,avgCount = 0,1
    newSheet.cell(row=4,column=1).value = f"Average expenditure by {householdType} is ${format(avg/avgCount,'.2f')} per quarter"

    # this call the function to enter the values to the report
    insertReportValues(newSheet,change,changeCount,indexChangeList,costChangeList)
    newWorkBook.save(path)
    print("Excel report Generated")


# This Function take user input to make report base on a group or subgroup of category people spend on
def selectExpenditureType():
    # This takes input for type of category
    options = ['Food', 'Alcohol and tobacco', 'Clothing and footwear', 'Housing', 'Contents and services', 'Health', 'Transport', 'Communication', 'Recreation and culture', 'Education', 'Miscellaneous', 'Interest']
    group = pyip.inputMenu(options,prompt="Select the group of expenditure type:\n", numbered = True)
    # This takes input for sub category if user wants it
    result = selectSubgroup(group)
    subgroup = group if result == "No subgroup" else result

    return group,subgroup

# This Function take user input to select subgroup of category of items people spend on
def selectSubgroup(group):

    # All of the list are generate using filtering prices using filter in excel and python script.
    if(group == "Food"):
        options = ['No subgroup','Fruit & veg', 'Meat', 'Grocery food', 'Soft drinks', 'Restaurant meals']

    elif(group == "Alcohol and tobacco"):
        options = ['No subgroup','Alcohol', 'Cigarettes and tobacco']

    elif(group == "Clothing and footwear"):
        options = ['No subgroup','Clothing', 'Footwear']

    elif(group == "Housing"):
        options = ['No subgroup','Rent', 'Property maintenance', 'Property rates', 'Household energy']

    elif(group == "Contents and services"):
        options = ['No subgroup','Furniture', 'Textiles', 'Appliances', 'Utensils', 'Tools', 'Other supplies and services']

    elif(group == "Health"):
        options = ['No subgroup','Medical products', 'Out-patient services', 'Hospital services']

    elif(group == "Transport"):
        options = ['No subgroup','Vehicles', 'Private transport supplies', 'Passenger transport']

    elif(group == "Communication"):
        options = ['No subgroup','Postal Services', 'Telecommunication equip', 'Telecommunication services']

    elif(group == "Recreation and culture"):
        options = ['No subgroup','Audio-visual', 'Major recreational equip', 'Other recreational equip', 'Rec & cultural services', 'Newspapers & books', 'Accommodation']

    elif(group == "Education"):
        options = ['No subgroup','Early childhood education', 'Primary/secondary education', 'Tertiary education', 'Other education']

    elif(group == "Miscellaneous"):
        options = ['No subgroup','Personal care', 'Personal effects', 'Insurance', 'Credit services', 'Miscellaneous services']

    else:
        options = ['No subgroup']

    subgroup = pyip.inputMenu(options, numbered = True) 
    return subgroup

# This function is used to set the data and generate report in new excel file base on year and spend category
def setExpenditureTypeReport():
    # calling function selectYear to get year range
    start,end = selectYear()

    # calling this function to create report based on given data
    group,subgroup = selectExpenditureType()

    # This create file name based on year and categories
    path = createFile(start,end,False,group,subgroup)
    newWorkBook = openpyxl.load_workbook(path)
    newSheet = newWorkBook.active
    check = sheet['G']
    rowNewFile = 11
    avg = change = 0
    avgCount= changeCount = 0
    indexChangeList = []
    costChangeList = []

    # This sets the heading of report base on the categories of report user wants
    if start == end:
        newSheet.cell(row=1,column=1).value = f"Report on Expenditure on {subgroup} of year {end}"
    else:
        newSheet.cell(row=1,column=1).value = f"Report on Expenditure on {subgroup} from year {start} to {end}"

    # calling this function to add report headings
    enterReportHeadings(newSheet)

    # This loop select the data base on year and categories and add it to the report
    j = 1
    for i in check:

        # this is to skip the first iteration which adds column name
        if j != 1:
            if(sheet.cell(row = j,column = 3).value == None):
                break
            else:
                # this handles error if string can't be converted to integer
                try:
                    year = int(sheet.cell(row = j,column = 3).value[0:4])
                except ValueError:
                    print("Cannot convert the value to integer")

            if i.value == subgroup and  year >= start and year <= end:
                rowNewFile+=1

                # this handles error if string can't be converted to integer
                try:
                    # this is to get average money spend by household
                    avg += int(sheet.cell(row = j, column = 9).value)
                    # This adds purchase cost into list
                    costChange = sheet.cell(row= j, column=9).value
                    costChangeList.append(int(costChange))
                except ValueError:
                    print("Cannot convert the value to integer")
                avgCount+=1

                for k in range(1,12):
                    c1 = sheet.cell(row=j, column=k)
                    newSheet.cell(row=rowNewFile, column=k).value = c1.value
                temp = sheet.cell(row = j, column = 11).value

                if(temp != None and temp != "NA"):
                    # This adds purchase index into list
                    indexChangeList.append(int(temp))
                    change += int(temp)
                    changeCount += 1
        j+=1

    # this is to prevent zero division error
    if avgCount == 0 : avg,avgCount = 0,1
    newSheet.cell(row=4,column=1).value = f"Average expenditure on {subgroup} is ${format(avg/avgCount,'.2f')} per quarter"
    # this call the function to enter the values to the report
    insertReportValues(newSheet,change,changeCount,indexChangeList,costChangeList)
    newWorkBook.save(path)
    print("Excel report Generated")


# This functions is use to make report dependent on both household and expenditure type
def selectBoth():
    start,end = selectYear()

    householdType = selectIncomeType()
    group,subgroup = selectExpenditureType()

    # This creates the file for this report
    path = createFile(start,end,householdType,group,subgroup)
    newWorkBook = openpyxl.load_workbook(path)
    newSheet = newWorkBook.active
    check = sheet['G']
    rowNewFile = 11
    avg = change = 0
    avgCount= changeCount = 0
    indexChangeList = []
    costChangeList = []

    # This sets the heading of report base on the categories of report user wants
    if start == end:
        newSheet.cell(row=1,column=1).value = f"Report on Expenditure on {subgroup} of year {end}"
    else:
        newSheet.cell(row=1,column=1).value = f"Report on Expenditure on {subgroup} from year {start} to {end}"

    # calling this function to add report headings
    enterReportHeadings(newSheet)

    # This loop select the data base on year and categories and add it to the report
    j = 1
    for i in check:
        # this is to skip the first iteration which adds column name
        if j != 1: 
            hht = sheet.cell(row = j,column = 1).value                   # household column selection
            if(sheet.cell(row = j,column = 3).value == None):
                break
            else:
                # this handles error if string can't be converted to integer
                try:
                    year = int(sheet.cell(row = j,column = 3).value[0:4]) # year column selection
                except ValueError:
                    print("Cannot convert the value to integer")   

            if i.value == subgroup and  year >= start and year <= end and hht == householdType:
                rowNewFile+=1
                # this handles error if string can't be converted to integer
                try:
                    # this is to get average money spend by household
                    avg += int(sheet.cell(row = j, column = 9).value)
                    # This adds purchase cost into list
                    costChange = sheet.cell(row= j, column=9).value
                    costChangeList.append(int(costChange))
                except ValueError:
                    print("Cannot convert the value to integer")   
                avgCount+=1

                for k in range(1,12):
                    c1 = sheet.cell(row=j, column=k)
                    newSheet.cell(row=rowNewFile, column=k).value = c1.value
                temp = sheet.cell(row = j, column = 11).value

                if(temp != None and temp != "NA"):
                    # This adds purchase index into list
                    indexChangeList.append(int(temp))
                    change += int(temp)
                    changeCount += 1
        j+=1

    # this is to prevent zero division error
    if avgCount == 0 : avg,avgCount = 0,1
    newSheet.cell(row=4,column=1).value = f"Average expenditure on {subgroup} is ${format(avg/avgCount,'.2f')} per quarter"
    
    # this call the function to enter the values to the report
    insertReportValues(newSheet,change,changeCount,indexChangeList,costChangeList)
    newWorkBook.save(path)
    print("Excel report Generated")

    

# This function set the calculated values into the report
def insertReportValues(newSheet,change,changeCount,indexChangeList,costChangeList):
    
    if changeCount == 0 : change,changeCount = 0,1
    newSheet.cell(row=5,column=1).value = f"Annul expenditure change is {format(100*(int(change)/changeCount),'.2f')}"
    newSheet.cell(row=6,column=1).value = f"Maximum index increase was {max(indexChangeList)}"
    newSheet.cell(row=7,column=1).value = f"Maximum index drop was {min(indexChangeList)}"
    newSheet.cell(row=8,column=1).value = f"Maximum Amount spent was ${max(costChangeList)}"
    newSheet.cell(row=9,column=1).value = f"Lowest Amount spent was ${min(costChangeList)}"
    

# This Function ask user for year inputS
def selectYear():
    print("Enter the range from year 2008 to 2023")
    print("For single year enter the year twice")
    start = pyip.inputInt(prompt="Enter the start year: ",min = 2008,max = 2023,blank=False)
    end = pyip.inputInt(prompt="Enter the end year: ",min = start,max = 2023,blank=False)
    
    return start,end

# This function will ask employee which type of report they wanted to generate
def selectReportType():
    reportType = pyip.inputMenu(["By Household type","By Expenditure type","By Both"],prompt="Choose the type of report you want to create\n", numbered = True)

    if(reportType == "By Household type" ):
        setIncomeTypeReport()
    elif(reportType == "By Expenditure type" ):
        setExpenditureTypeReport()
    elif(reportType == "By Both"):
        selectBoth()
    
# This function call executes the report generation  
selectReportType()